# plantilla_excel.py
# Lee el Excel "modelo" de la ficha (el mismo que usa el ayuntamiento para
# imprimirlas a mano) y extrae de ahi el municipio/nucleo/provincia y el
# escudo, para no tener que escribirlos a mano en el movil cada vez. Se
# prepara UNA vez en el ordenador (cambiando el logo/nombre en el Excel) y
# solo hace falta elegir ese archivo desde el movil.
#
# Usa openpyxl, que es Python puro (sin extensiones compiladas en C como
# reportlab o fpdf2), asi que no deberia tener los problemas de
# compilacion para Android que tuvimos con esas otras librerias.
#
# Las celdas C3/C4/C5 (municipio/nucleo/provincia) y la posicion del
# escudo (columnas izquierdas, frente al logo SOMACyL que va a la
# derecha) coinciden con la estructura real de la ficha modelo del
# cliente (comprobado con Ficha_Modelo_Valle_de_Valdelaguna.xlsx).

import openpyxl


def leer_plantilla_excel(ruta_xlsx):
    """Devuelve {"municipio", "nucleo", "provincia", "escudo_bytes",
    "escudo_ext"} leidos del Excel. escudo_bytes es None si no se
    encontro ninguna imagen en las columnas de la izquierda (donde va el
    escudo en la plantilla). Lanza excepcion si el archivo no se puede
    abrir como Excel (formato inesperado)."""
    wb = openpyxl.load_workbook(ruta_xlsx, data_only=True)
    ws = wb.active

    municipio = str(ws["C3"].value or "").strip()
    nucleo = str(ws["C4"].value or "").strip()
    provincia = str(ws["C5"].value or "").strip()

    # La celda del Excel a veces ya trae el texto completo "AYUNTAMIENTO DE
    # ...", pero generar_ficha_pdf.py añade ese prefijo por su cuenta al
    # dibujar la cabecera -> se quita aqui para no duplicarlo.
    prefijo = "ayuntamiento de"
    if municipio.lower().startswith(prefijo):
        municipio = municipio[len(prefijo):].strip()

    escudo_bytes = None
    escudo_ext = "png"
    for img in getattr(ws, "_images", []):
        try:
            col_inicial = img.anchor._from.col
        except Exception:
            col_inicial = 99
        # El escudo esta anclado en las columnas de la izquierda (donde
        # esta en la plantilla original); el logo SOMACyL (que no cambia
        # de un municipio a otro) esta mas a la derecha y se ignora aqui.
        if col_inicial <= 3:
            datos = img.ref.getvalue() if hasattr(img.ref, "getvalue") else None
            if datos:
                escudo_bytes = datos
                escudo_ext = "png" if datos[:8] == b"\x89PNG\r\n\x1a\n" else "jpg"
                break

    return {
        "municipio": municipio,
        "nucleo": nucleo,
        "provincia": provincia,
        "escudo_bytes": escudo_bytes,
        "escudo_ext": escudo_ext,
    }
